"""
Bot de Telegram para buscar expedientes laborales — con menú de botones.

Roles:
  - admin: acceso total + gestión de usuarios.
  - editor: puede buscar, agregar, dar de baja, retirar y eliminar expedientes.
  - lector: solo puede buscar y ver el listado/exportar.

El menú de botones (fijo abajo) cambia según el rol de quien escribe.
Casi no hace falta escribir comandos: solo tocar los botones. Las únicas
veces que se escribe texto libre es para indicar el nombre, carnet o
ubicación de un expediente (eso no se puede reemplazar por botones).

Autor: generado con Claude
"""

import logging
import os
import re
import sqlite3
from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    ConversationHandler,
    CallbackQueryHandler,
    filters,
)

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

_DB_DIR = os.environ.get("DB_DIR", os.path.dirname(__file__))
os.makedirs(_DB_DIR, exist_ok=True)
DB_PATH = os.path.join(_DB_DIR, "expedientes.db")

# ID de Telegram del administrador (tu cuenta). Siempre tiene acceso total.
ADMIN_USER_ID = 1186207945

DEPARTAMENTOS = [
    "Dirección General",
    "Dirección de Gestión y Desarrollo Organizacional",
    "Dirección Contable Financiera",
    "Dirección de Capital Humano",
    "Dirección de Logística",
    "Dirección Comercial",
    "Dirección de Comercio Exterior",
    "División de Sistemas Informáticos",
    "División de Servicios Técnicos",
    "División de Infraestructura y Comunicaciones",
    "División de Ciberseguridad",
]

ESTADOS_EXPEDIENTE = ["Activo", "Baja"]

# ---------- Lenguaje natural ----------

NOMBRE_BOT = "Lucas"

PATRONES_BUSQUEDA = [
    r"d[oó]nde est[aá]\s+(?:el\s+)?expediente\s+de\s+(.+)",
    r"d[oó]nde est[aá]\s+(.+)",
    r"busca(?:r)?\s+(?:a\s+|el\s+expediente\s+de\s+)?(.+)",
    r"expediente\s+de\s+(.+)",
    r"ubicaci[oó]n\s+de\s+(.+)",
]


def detectar_busqueda_natural(texto):
    """
    Si el texto parece una pregunta/orden de búsqueda en lenguaje natural,
    devuelve el término a buscar (nombre o carnet). Si no, devuelve None.
    """
    texto_limpio = re.sub(rf"^{NOMBRE_BOT}[,:]?\s*", "", texto.strip(), flags=re.IGNORECASE)
    texto_lower = texto_limpio.lower().strip(" ?¿!¡.")

    for patron in PATRONES_BUSQUEDA:
        match = re.search(patron, texto_lower, flags=re.IGNORECASE)
        if match:
            termino = match.group(1).strip(" ?¿!¡.")
            if termino:
                return termino
    return None


# ---------- Textos de los botones del menú ----------
BTN_BUSCAR = "🔍 Buscar expediente"
BTN_AGREGAR = "➕ Agregar expediente"
BTN_BAJA = "⬇️ Marcar Baja"
BTN_RETIRAR = "📤 Marcar Retirado"
BTN_ELIMINAR = "🗑️ Eliminar expediente"
BTN_EDITAR = "✏️ Editar expediente"
BTN_LISTA = "📊 Resumen"
BTN_EXPORTAR = "📁 Exportar a Excel"
BTN_USUARIOS = "👥 Usuarios"
BTN_AYUDA = "❓ Ayuda"
BTN_CANCELAR = "✖️ Cancelar"


# ---------- Base de datos ----------

def init_db():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS expedientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            carnet TEXT,
            estado TEXT NOT NULL DEFAULT 'Activo',
            departamento TEXT,
            ubicacion TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS usuarios_autorizados (
            user_id INTEGER PRIMARY KEY,
            nombre TEXT,
            username TEXT,
            rol TEXT NOT NULL DEFAULT 'editor'
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS solicitudes_pendientes (
            user_id INTEGER PRIMARY KEY,
            nombre TEXT,
            username TEXT
        )
        """
    )
    conn.commit()
    conn.close()


def es_admin(user_id):
    return user_id == ADMIN_USER_ID


def esta_autorizado(user_id):
    if es_admin(user_id):
        return True
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM usuarios_autorizados WHERE user_id = ?", (user_id,))
    row = cur.fetchone()
    conn.close()
    return row is not None


def obtener_rol(user_id):
    if es_admin(user_id):
        return "admin"
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT rol FROM usuarios_autorizados WHERE user_id = ?", (user_id,))
    row = cur.fetchone()
    conn.close()
    return row[0] if row else None


def puede_editar(user_id):
    return obtener_rol(user_id) in ("admin", "editor")


def hay_solicitud_pendiente(user_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM solicitudes_pendientes WHERE user_id = ?", (user_id,))
    row = cur.fetchone()
    conn.close()
    return row is not None


def crear_solicitud(user_id, nombre, username):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        "INSERT OR REPLACE INTO solicitudes_pendientes (user_id, nombre, username) VALUES (?, ?, ?)",
        (user_id, nombre, username),
    )
    conn.commit()
    conn.close()


def eliminar_solicitud(user_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("DELETE FROM solicitudes_pendientes WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()


def aprobar_usuario(user_id, nombre, username, rol="editor"):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        "INSERT OR REPLACE INTO usuarios_autorizados (user_id, nombre, username, rol) VALUES (?, ?, ?, ?)",
        (user_id, nombre, username, rol),
    )
    conn.commit()
    conn.close()


def revocar_usuario(user_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("DELETE FROM usuarios_autorizados WHERE user_id = ?", (user_id,))
    affected = cur.rowcount
    conn.commit()
    conn.close()
    return affected


def listar_usuarios_autorizados():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT user_id, nombre, username, rol FROM usuarios_autorizados")
    rows = cur.fetchall()
    conn.close()
    return rows


def formatear_expediente(nombre, carnet, estado, departamento, ubicacion):
    return (
        f"👤 {nombre}\n"
        f"🪪 Carnet: {carnet or '—'}\n"
        f"📌 Estado: {estado}\n"
        f"🏢 Departamento: {departamento or '—'}\n"
        f"📍 Ubicación: {ubicacion}"
    )


async def responder_busqueda(update: Update, termino: str):
    """Busca el término y responde: nada, un resultado completo, o botones para elegir si hay varios."""
    resultados = buscar_expediente_con_id(termino)

    if not resultados:
        await update.message.reply_text(
            f"No encontré ningún expediente que coincida con '{termino}'.",
            reply_markup=construir_menu(update.effective_user.id),
        )
        return

    if len(resultados) == 1:
        expediente_id, nombre, carnet, estado, departamento, ubicacion = resultados[0]
        await update.message.reply_text(
            formatear_expediente(nombre, carnet, estado, departamento, ubicacion),
            reply_markup=construir_menu(update.effective_user.id),
        )
        return

    # Varios resultados: mostrar botones para elegir
    filas = []
    for expediente_id, nombre, carnet, estado, departamento, ubicacion in resultados:
        etiqueta = f"{nombre} — {carnet or 'sin carnet'} ({estado})"
        filas.append([InlineKeyboardButton(etiqueta, callback_data=f"ver_exp:{expediente_id}")])
    await update.message.reply_text(
        f"Encontré {len(resultados)} expedientes que coinciden con '{termino}'. ¿Cuál quieres ver?",
        reply_markup=InlineKeyboardMarkup(filas),
    )


async def manejar_ver_expediente(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    expediente_id = int(query.data.split(":")[1])
    row = obtener_expediente_por_id(expediente_id)
    if not row:
        await query.edit_message_text("Ese expediente ya no existe.")
        return
    _, nombre, carnet, estado, departamento, ubicacion = row
    await query.edit_message_text(formatear_expediente(nombre, carnet, estado, departamento, ubicacion))


def buscar_expediente_con_id(termino):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    like = f"%{termino}%"
    cur.execute(
        """
        SELECT id, nombre, carnet, estado, departamento, ubicacion
        FROM expedientes
        WHERE nombre LIKE ? OR carnet LIKE ?
        """,
        (like, like),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def buscar_por_carnet_exacto(carnet):
    """Busca un expediente cuyo carnet coincida EXACTAMENTE (ignora '-' o vacío)."""
    if not carnet or carnet.strip() in ("-", ""):
        return None
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        "SELECT id, nombre, carnet, estado, departamento, ubicacion FROM expedientes WHERE carnet = ?",
        (carnet.strip(),),
    )
    row = cur.fetchone()
    conn.close()
    return row


def obtener_expediente_por_id(expediente_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        "SELECT id, nombre, carnet, estado, departamento, ubicacion FROM expedientes WHERE id = ?",
        (expediente_id,),
    )
    row = cur.fetchone()
    conn.close()
    return row


def actualizar_campo_expediente(expediente_id, campo, valor):
    columnas_validas = {"nombre", "carnet", "estado", "departamento", "ubicacion"}
    if campo not in columnas_validas:
        raise ValueError("Campo no válido")
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(f"UPDATE expedientes SET {campo} = ? WHERE id = ?", (valor, expediente_id))
    conn.commit()
    conn.close()


def actualizar_expediente(expediente_id, estado, departamento, ubicacion):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        "UPDATE expedientes SET estado = ?, departamento = ?, ubicacion = ? WHERE id = ?",
        (estado, departamento, ubicacion, expediente_id),
    )
    conn.commit()
    conn.close()


def agregar_expediente(nombre, carnet, estado, departamento, ubicacion):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO expedientes (nombre, carnet, estado, departamento, ubicacion)
        VALUES (?, ?, ?, ?, ?)
        """,
        (nombre, carnet, estado, departamento, ubicacion),
    )
    conn.commit()
    conn.close()


def marcar_baja(termino):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    like = f"%{termino}%"
    cur.execute(
        "UPDATE expedientes SET estado = 'Baja' WHERE nombre LIKE ? OR carnet LIKE ?",
        (like, like),
    )
    affected = cur.rowcount
    conn.commit()
    conn.close()
    return affected


def marcar_retirado(termino):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    like = f"%{termino}%"
    cur.execute(
        "UPDATE expedientes SET estado = 'Retirado' WHERE nombre LIKE ? OR carnet LIKE ?",
        (like, like),
    )
    affected = cur.rowcount
    conn.commit()
    conn.close()
    return affected


def eliminar_expediente_por_id(expediente_id):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("DELETE FROM expedientes WHERE id = ?", (expediente_id,))
    affected = cur.rowcount
    conn.commit()
    conn.close()
    return affected


def contar_expedientes():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT estado, COUNT(*) FROM expedientes GROUP BY estado")
    rows = cur.fetchall()
    conn.close()
    return rows


def obtener_todos_expedientes():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        "SELECT nombre, carnet, estado, departamento, ubicacion FROM expedientes ORDER BY nombre"
    )
    rows = cur.fetchall()
    conn.close()
    return rows


# ---------- Menú según rol ----------

def construir_menu(user_id):
    rol = obtener_rol(user_id)
    filas = [[BTN_BUSCAR], [BTN_LISTA, BTN_EXPORTAR]]

    if rol in ("admin", "editor"):
        filas.insert(1, [BTN_AGREGAR])
        filas.append([BTN_BAJA, BTN_RETIRAR])
        filas.append([BTN_ELIMINAR])
        filas.append([BTN_EDITAR])

    if rol == "admin":
        filas.append([BTN_USUARIOS])

    filas.append([BTN_AYUDA])
    return ReplyKeyboardMarkup(filas, resize_keyboard=True)


def construir_teclado_estados():
    return InlineKeyboardMarkup(
        [[InlineKeyboardButton(e, callback_data=f"estado:{e}")] for e in ESTADOS_EXPEDIENTE]
    )


def construir_teclado_departamentos():
    filas = [[InlineKeyboardButton(d, callback_data=f"depto:{i}")] for i, d in enumerate(DEPARTAMENTOS)]
    return InlineKeyboardMarkup(filas)


# ---------- Control de acceso ----------

async def verificar_acceso(update: Update) -> bool:
    user = update.effective_user
    if esta_autorizado(user.id):
        return True

    if hay_solicitud_pendiente(user.id):
        await update.message.reply_text(
            "Tu solicitud de acceso ya fue enviada y está esperando aprobación."
        )
        return False

    crear_solicitud(user.id, user.full_name, user.username or "")
    await update.message.reply_text(
        "No tienes acceso a este bot todavía. Se envió una solicitud al administrador.\n"
        "Te avisaré aquí mismo cuando sea aprobada o rechazada."
    )

    teclado = InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("✅ Aprobar (Editor)", callback_data=f"aprobar:editor:{user.id}")],
            [InlineKeyboardButton("🔍 Aprobar (Solo lectura)", callback_data=f"aprobar:lector:{user.id}")],
            [InlineKeyboardButton("❌ Rechazar", callback_data=f"rechazar:-:{user.id}")],
        ]
    )
    username_txt = f"@{user.username}" if user.username else "(sin username)"
    await update.get_bot().send_message(
        chat_id=ADMIN_USER_ID,
        text=(
            "🔔 Nueva solicitud de acceso al bot:\n\n"
            f"👤 {user.full_name} {username_txt}\n"
            f"🆔 ID: {user.id}"
        ),
        reply_markup=teclado,
    )
    return False


async def manejar_aprobacion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.from_user.id != ADMIN_USER_ID:
        await query.answer("Solo el administrador puede hacer esto.", show_alert=True)
        return

    accion, rol, user_id_str = query.data.split(":")
    user_id = int(user_id_str)

    if accion == "aprobar":
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT nombre, username FROM solicitudes_pendientes WHERE user_id = ?", (user_id,))
        row = cur.fetchone()
        conn.close()
        nombre, username = row if row else ("", "")

        aprobar_usuario(user_id, nombre, username, rol)
        eliminar_solicitud(user_id)
        rol_legible = "Editor (puede agregar y modificar)" if rol == "editor" else "Solo lectura (solo puede buscar)"
        await query.edit_message_text(f"✅ Aprobado: {nombre} (ID {user_id})\nRol: {rol_legible}")
        await context.bot.send_message(
            chat_id=user_id,
            text="✅ Tu acceso fue aprobado. Ya puedes usar el bot, escribe /start.",
        )
    else:
        eliminar_solicitud(user_id)
        await query.edit_message_text(f"❌ Rechazado: ID {user_id}")
        await context.bot.send_message(
            chat_id=user_id,
            text="❌ Tu solicitud de acceso fue rechazada.",
        )


# ---------- Comandos básicos / menú ----------

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await verificar_acceso(update):
        return
    await update.message.reply_text(
        f"Hola, soy {NOMBRE_BOT} 🤖, tu asistente de expedientes laborales.\n\n"
        "Puedes tocar una opción del menú de abajo, o simplemente escribirme algo como:\n"
        f"\"{NOMBRE_BOT}, ¿dónde está el expediente de Ali?\"",
        reply_markup=construir_menu(update.effective_user.id),
    )


async def ayuda(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await verificar_acceso(update):
        return
    texto = (
        f"Soy {NOMBRE_BOT}. Puedes usar los botones del menú, o escribirme directamente, por ejemplo:\n"
        f"\"{NOMBRE_BOT}, ¿dónde está el expediente de Ali?\"\n\n"
        f"{BTN_BUSCAR} - busca un expediente por nombre o carnet\n"
        f"{BTN_LISTA} - resumen de cuántos expedientes hay\n"
        f"{BTN_EXPORTAR} - descarga el listado completo en Excel\n"
    )
    if puede_editar(update.effective_user.id):
        texto += (
            f"{BTN_AGREGAR} - agrega un expediente nuevo\n"
            f"{BTN_BAJA} - marca un expediente como Baja\n"
            f"{BTN_RETIRAR} - marca un expediente como Retirado (se lo llevaron)\n"
            f"{BTN_ELIMINAR} - borra un expediente para siempre\n"
            f"{BTN_EDITAR} - cambia nombre, carnet, ubicación, departamento o estado de un expediente\n"
        )
    if es_admin(update.effective_user.id):
        texto += (
            f"{BTN_USUARIOS} - lista usuarios con acceso\n"
            "/rol <id> <editor|lector> - cambia el rol de un usuario\n"
            "/revocar <id> - quita el acceso a un usuario\n"
        )
    await update.message.reply_text(texto, reply_markup=construir_menu(update.effective_user.id))


async def mostrar_lista(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await verificar_acceso(update):
        return
    rows = contar_expedientes()
    if not rows:
        await update.message.reply_text("Todavía no hay expedientes registrados.")
        return
    texto = "Resumen de expedientes:\n\n"
    for estado, cantidad in rows:
        texto += f"{estado}: {cantidad}\n"
    await update.message.reply_text(texto)


async def mostrar_usuarios(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not es_admin(update.effective_user.id):
        await update.message.reply_text("Solo el administrador puede usar esta opción.")
        return
    rows = listar_usuarios_autorizados()
    if not rows:
        await update.message.reply_text("No hay usuarios adicionales con acceso (solo tú).")
        return
    texto = "Usuarios con acceso:\n\n"
    for user_id, nombre, username, rol in rows:
        username_txt = f"@{username}" if username else "(sin username)"
        rol_txt = "✏️ Editor" if rol == "editor" else "🔍 Solo lectura"
        texto += f"🆔 {user_id} - {nombre} {username_txt} - {rol_txt}\n"
    texto += "\nPara cambiar un rol: /rol <id> <editor|lector>\nPara quitar acceso: /revocar <id>"
    await update.message.reply_text(texto)


async def cambiar_rol(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not es_admin(update.effective_user.id):
        await update.message.reply_text("Solo el administrador puede usar este comando.")
        return
    if len(context.args) != 2 or context.args[1] not in ("editor", "lector"):
        await update.message.reply_text("Escribe así: /rol 123456789 editor   (o /rol 123456789 lector)")
        return
    try:
        user_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("Ese ID no es válido.")
        return
    nuevo_rol = context.args[1]
    if not esta_autorizado(user_id) or es_admin(user_id):
        await update.message.reply_text("Ese ID no está en la lista de usuarios autorizados.")
        return
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("UPDATE usuarios_autorizados SET rol = ? WHERE user_id = ?", (nuevo_rol, user_id))
    conn.commit()
    conn.close()
    await update.message.reply_text(f"Rol actualizado para el ID {user_id}: {nuevo_rol}")


async def revocar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not es_admin(update.effective_user.id):
        await update.message.reply_text("Solo el administrador puede usar este comando.")
        return
    if not context.args:
        await update.message.reply_text("Escribe así: /revocar 123456789")
        return
    try:
        user_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("Ese ID no es válido.")
        return
    afectados = revocar_usuario(user_id)
    if afectados:
        await update.message.reply_text(f"Acceso revocado para el ID {user_id}.")
    else:
        await update.message.reply_text("Ese ID no estaba en la lista de usuarios autorizados.")


async def exportar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await verificar_acceso(update):
        return
    filas = obtener_todos_expedientes()
    if not filas:
        await update.message.reply_text("Todavía no hay expedientes registrados para exportar.")
        return

    await update.message.reply_text("Generando el archivo Excel, un momento...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Expedientes"

    encabezados = ["Nombre", "Carnet de Identidad", "Estado", "Departamento/Dirección", "Ubicación"]
    ws.append(encabezados)
    for fila in filas:
        ws.append(list(fila))

    num_filas = len(filas) + 1
    num_columnas = len(encabezados)
    ultima_columna = get_column_letter(num_columnas)
    rango_tabla = f"A1:{ultima_columna}{num_filas}"

    tabla = Table(displayName="Expedientes", ref=rango_tabla)
    estilo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)

    for col_idx in range(1, num_columnas + 1):
        letra = get_column_letter(col_idx)
        max_len = max(
            [len(str(encabezados[col_idx - 1]))]
            + [len(str(fila[col_idx - 1])) for fila in filas if fila[col_idx - 1] is not None]
        )
        ws.column_dimensions[letra].width = max_len + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    await update.message.reply_document(
        document=buffer,
        filename="expedientes.xlsx",
        caption=f"📊 Listado completo: {len(filas)} expediente(s).",
    )


async def cancelar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text(
        "Operación cancelada.", reply_markup=construir_menu(update.effective_user.id)
    )
    return ConversationHandler.END


# ---------- Flujo: Buscar ----------

ESPERANDO_BUSQUEDA = 100


async def buscar_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await verificar_acceso(update):
        return ConversationHandler.END
    await update.message.reply_text(
        "Escribe el nombre o el carnet de identidad que quieres buscar.",
        reply_markup=ReplyKeyboardMarkup([[BTN_CANCELAR]], resize_keyboard=True),
    )
    return ESPERANDO_BUSQUEDA


async def buscar_recibir(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await responder_busqueda(update, update.message.text)
    return ConversationHandler.END


# ---------- Flujo: Agregar ----------

NOMBRE, CARNET, ESTADO, DEPARTAMENTO, UBICACION, DECIDIR_DUPLICADO = range(101, 107)


async def agregar_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await verificar_acceso(update):
        return ConversationHandler.END
    if not puede_editar(update.effective_user.id):
        await update.message.reply_text("No tienes permiso para agregar expedientes (solo puedes buscar).")
        return ConversationHandler.END
    await update.message.reply_text(
        "Vamos a agregar un expediente nuevo.\n\n¿Cuál es el nombre completo?",
        reply_markup=ReplyKeyboardMarkup([[BTN_CANCELAR]], resize_keyboard=True),
    )
    return NOMBRE


async def agregar_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["nombre"] = update.message.text
    await update.message.reply_text("¿Cuál es el carnet de identidad? (o escribe - si no aplica)")
    return CARNET


async def agregar_carnet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    carnet = update.message.text.strip()
    context.user_data["carnet"] = carnet

    existente = buscar_por_carnet_exacto(carnet)
    if existente:
        expediente_id, nombre, carnet_existente, estado, departamento, ubicacion = existente
        context.user_data["expediente_existente_id"] = expediente_id
        teclado = InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("🔄 Actualizar este expediente", callback_data="usar_existente")],
                [InlineKeyboardButton("➕ Crear uno nuevo de todas formas", callback_data="crear_nuevo")],
            ]
        )
        await update.message.reply_text(
            "⚠️ Ya existe un expediente con ese carnet de identidad:\n\n"
            f"{formatear_expediente(nombre, carnet_existente, estado, departamento, ubicacion)}\n\n"
            "¿Qué quieres hacer? (por ejemplo, si esta persona se está recontratando, "
            "lo más práctico es actualizar el expediente existente con el estado y ubicación nuevos)",
            reply_markup=teclado,
        )
        return DECIDIR_DUPLICADO

    await update.message.reply_text(
        "Selecciona el estado:", reply_markup=construir_teclado_estados()
    )
    return ESTADO


async def decidir_duplicado(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "crear_nuevo":
        context.user_data.pop("expediente_existente_id", None)
        await query.edit_message_text("De acuerdo, vamos a crear un expediente nuevo (separado).")
    else:
        await query.edit_message_text("De acuerdo, vamos a actualizar el expediente existente.")

    await query.message.reply_text(
        "Selecciona el estado:", reply_markup=construir_teclado_estados()
    )
    return ESTADO


async def agregar_estado_boton(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    estado = query.data.split(":")[1]
    context.user_data["estado"] = estado
    await query.edit_message_text(f"Estado seleccionado: {estado}")
    await query.message.reply_text(
        "Selecciona el departamento o dirección:", reply_markup=construir_teclado_departamentos()
    )
    return DEPARTAMENTO


async def agregar_departamento_boton(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    indice = int(query.data.split(":")[1])
    departamento = DEPARTAMENTOS[indice]
    context.user_data["departamento"] = departamento
    await query.edit_message_text(f"Departamento seleccionado: {departamento}")
    await query.message.reply_text("¿Cuál es la ubicación física del expediente? (ej. ABT-12, Estante 3)")
    return UBICACION


async def agregar_ubicacion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["ubicacion"] = update.message.text
    d = context.user_data

    expediente_existente_id = d.get("expediente_existente_id")
    if expediente_existente_id:
        actualizar_expediente(expediente_existente_id, d["estado"], d["departamento"], d["ubicacion"])
        await update.message.reply_text(
            f"✅ Expediente actualizado (se recontrató / reingresó):\n\n"
            f"👤 {d['nombre']}\n"
            f"🪪 Carnet: {d['carnet']}\n"
            f"📌 Estado: {d['estado']}\n"
            f"🏢 Departamento: {d['departamento']}\n"
            f"📍 Ubicación: {d['ubicacion']}",
            reply_markup=construir_menu(update.effective_user.id),
        )
    else:
        agregar_expediente(d["nombre"], d["carnet"], d["estado"], d["departamento"], d["ubicacion"])
        await update.message.reply_text(
            f"✅ Expediente guardado:\n\n"
            f"👤 {d['nombre']}\n"
            f"🪪 Carnet: {d['carnet']}\n"
            f"📌 Estado: {d['estado']}\n"
            f"🏢 Departamento: {d['departamento']}\n"
            f"📍 Ubicación: {d['ubicacion']}",
            reply_markup=construir_menu(update.effective_user.id),
        )
    context.user_data.clear()
    return ConversationHandler.END


# ---------- Flujo: Baja / Retirar / Eliminar (piden nombre o carnet) ----------

ESPERANDO_BAJA, ESPERANDO_RETIRAR, ESPERANDO_ELIMINAR = 201, 202, 203


async def baja_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await verificar_acceso(update):
        return ConversationHandler.END
    if not puede_editar(update.effective_user.id):
        await update.message.reply_text("No tienes permiso para modificar expedientes (solo puedes buscar).")
        return ConversationHandler.END
    await update.message.reply_text(
        "Escribe el nombre o carnet del expediente que quieres marcar como Baja.",
        reply_markup=ReplyKeyboardMarkup([[BTN_CANCELAR]], resize_keyboard=True),
    )
    return ESPERANDO_BAJA


async def baja_recibir(update: Update, context: ContextTypes.DEFAULT_TYPE):
    termino = update.message.text
    afectados = marcar_baja(termino)
    if afectados:
        texto = f"Se marcó como Baja: {afectados} expediente(s)."
    else:
        texto = f"No encontré ningún expediente que coincida con '{termino}'."
    await update.message.reply_text(texto, reply_markup=construir_menu(update.effective_user.id))
    return ConversationHandler.END


async def retirar_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await verificar_acceso(update):
        return ConversationHandler.END
    if not puede_editar(update.effective_user.id):
        await update.message.reply_text("No tienes permiso para modificar expedientes (solo puedes buscar).")
        return ConversationHandler.END
    await update.message.reply_text(
        "Escribe el nombre o carnet del expediente que se llevaron (se marcará como Retirado).",
        reply_markup=ReplyKeyboardMarkup([[BTN_CANCELAR]], resize_keyboard=True),
    )
    return ESPERANDO_RETIRAR


async def retirar_recibir(update: Update, context: ContextTypes.DEFAULT_TYPE):
    termino = update.message.text
    afectados = marcar_retirado(termino)
    if afectados:
        texto = (
            f"📤 Se marcó como Retirado: {afectados} expediente(s).\n"
            "El registro se conserva en la base de datos para historial."
        )
    else:
        texto = f"No encontré ningún expediente que coincida con '{termino}'."
    await update.message.reply_text(texto, reply_markup=construir_menu(update.effective_user.id))
    return ConversationHandler.END


async def eliminar_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await verificar_acceso(update):
        return ConversationHandler.END
    if not puede_editar(update.effective_user.id):
        await update.message.reply_text("No tienes permiso para modificar expedientes (solo puedes buscar).")
        return ConversationHandler.END
    await update.message.reply_text(
        "Escribe el nombre o carnet del expediente que quieres ELIMINAR PERMANENTEMENTE.",
        reply_markup=ReplyKeyboardMarkup([[BTN_CANCELAR]], resize_keyboard=True),
    )
    return ESPERANDO_ELIMINAR


async def eliminar_recibir(update: Update, context: ContextTypes.DEFAULT_TYPE):
    termino = update.message.text
    resultados = buscar_expediente_con_id(termino)
    if not resultados:
        await update.message.reply_text(
            f"No encontré ningún expediente que coincida con '{termino}'.",
            reply_markup=construir_menu(update.effective_user.id),
        )
        return ConversationHandler.END
    if len(resultados) > 1:
        await update.message.reply_text(
            "Encontré más de un expediente con ese término. Sé más específico (usa el carnet).",
            reply_markup=construir_menu(update.effective_user.id),
        )
        return ConversationHandler.END

    expediente_id, nombre, carnet, estado, departamento, ubicacion = resultados[0]
    teclado = InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton("🗑️ Sí, eliminar", callback_data=f"confirmar_eliminar:{expediente_id}"),
                InlineKeyboardButton("Cancelar", callback_data="cancelar_eliminar"),
            ]
        ]
    )
    await update.message.reply_text(
        f"⚠️ ¿Seguro que quieres ELIMINAR PERMANENTEMENTE este expediente?\n\n"
        f"👤 {nombre}\n🪪 Carnet: {carnet or '—'}\n📍 Ubicación: {ubicacion}\n\n"
        "Esta acción no se puede deshacer.",
        reply_markup=teclado,
    )
    await update.message.reply_text(
        "Mientras decides, puedes volver al menú:", reply_markup=construir_menu(update.effective_user.id)
    )
    return ConversationHandler.END


async def manejar_confirmacion_eliminar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if not puede_editar(query.from_user.id):
        await query.edit_message_text("No tienes permiso para hacer esto.")
        return

    if query.data == "cancelar_eliminar":
        await query.edit_message_text("Eliminación cancelada.")
        return

    expediente_id = int(query.data.split(":")[1])
    afectados = eliminar_expediente_por_id(expediente_id)
    if afectados:
        await query.edit_message_text("🗑️ Expediente eliminado permanentemente.")
    else:
        await query.edit_message_text("Ese expediente ya no existe (puede que se haya eliminado antes).")


# ---------- Flujo: Editar expediente ----------

ESPERANDO_BUSQUEDA_EDITAR, ELIGIENDO_CAMPO, ESPERANDO_VALOR_CAMPO = 301, 302, 303

CAMPOS_EDITABLES = {
    "nombre": "Nombre",
    "carnet": "Carnet de identidad",
    "ubicacion": "Ubicación",
}


async def editar_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await verificar_acceso(update):
        return ConversationHandler.END
    if not puede_editar(update.effective_user.id):
        await update.message.reply_text("No tienes permiso para editar expedientes (solo puedes buscar).")
        return ConversationHandler.END
    await update.message.reply_text(
        "Escribe el nombre o carnet del expediente que quieres editar.",
        reply_markup=ReplyKeyboardMarkup([[BTN_CANCELAR]], resize_keyboard=True),
    )
    return ESPERANDO_BUSQUEDA_EDITAR


async def editar_buscar_recibir(update: Update, context: ContextTypes.DEFAULT_TYPE):
    termino = update.message.text
    resultados = buscar_expediente_con_id(termino)

    if not resultados:
        await update.message.reply_text(
            f"No encontré ningún expediente que coincida con '{termino}'.",
            reply_markup=construir_menu(update.effective_user.id),
        )
        return ConversationHandler.END

    if len(resultados) > 1:
        filas = [
            [InlineKeyboardButton(f"{n} — {c or 'sin carnet'} ({e})", callback_data=f"editar_elegir:{i}")]
            for i, n, c, e, dep, u in resultados
        ]
        context.user_data["candidatos_editar"] = resultados
        await update.message.reply_text(
            "Encontré varios expedientes. ¿Cuál quieres editar?",
            reply_markup=InlineKeyboardMarkup(filas),
        )
        return ESPERANDO_BUSQUEDA_EDITAR

    return await iniciar_edicion_campo(update, context, resultados[0])


async def editar_elegir_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    expediente_id = int(query.data.split(":")[1])
    row = obtener_expediente_por_id(expediente_id)
    if not row:
        await query.edit_message_text("Ese expediente ya no existe.")
        return ConversationHandler.END
    await query.edit_message_text(f"Editando a: {row[1]}")
    return await iniciar_edicion_campo(update, context, row, es_callback=True)


async def iniciar_edicion_campo(update, context, expediente_row, es_callback=False):
    expediente_id, nombre, carnet, estado, departamento, ubicacion = expediente_row
    context.user_data["editar_id"] = expediente_id

    filas = [
        [InlineKeyboardButton("📍 Ubicación", callback_data="campo:ubicacion")],
        [InlineKeyboardButton("🏢 Departamento", callback_data="campo:departamento")],
        [InlineKeyboardButton("📌 Estado", callback_data="campo:estado")],
        [InlineKeyboardButton("🪪 Carnet", callback_data="campo:carnet")],
        [InlineKeyboardButton("👤 Nombre", callback_data="campo:nombre")],
    ]
    texto = (
        f"Expediente actual:\n\n{formatear_expediente(nombre, carnet, estado, departamento, ubicacion)}\n\n"
        "¿Qué campo quieres modificar?"
    )
    if es_callback:
        await update.callback_query.message.reply_text(texto, reply_markup=InlineKeyboardMarkup(filas))
    else:
        await update.message.reply_text(texto, reply_markup=InlineKeyboardMarkup(filas))
    return ELIGIENDO_CAMPO


async def editar_elegir_campo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    campo = query.data.split(":")[1]
    context.user_data["editar_campo"] = campo

    if campo == "estado":
        await query.edit_message_text("Selecciona el nuevo estado:")
        await query.message.reply_text(
            "Estado nuevo:",
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton(e, callback_data=f"nuevoestado:{e}")] for e in ["Activo", "Baja", "Retirado"]]
            ),
        )
        return ESPERANDO_VALOR_CAMPO

    if campo == "departamento":
        await query.edit_message_text("Selecciona el nuevo departamento:")
        await query.message.reply_text(
            "Departamento nuevo:", reply_markup=construir_teclado_departamentos()
        )
        return ESPERANDO_VALOR_CAMPO

    nombre_campo = CAMPOS_EDITABLES[campo]
    await query.edit_message_text(f"Escribe el nuevo valor para: {nombre_campo}")
    return ESPERANDO_VALOR_CAMPO


async def editar_guardar_valor_texto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    campo = context.user_data.get("editar_campo")
    expediente_id = context.user_data.get("editar_id")
    nuevo_valor = update.message.text.strip()

    actualizar_campo_expediente(expediente_id, campo, nuevo_valor)
    row = obtener_expediente_por_id(expediente_id)
    _, nombre, carnet, estado, departamento, ubicacion = row
    await update.message.reply_text(
        f"✅ Actualizado.\n\n{formatear_expediente(nombre, carnet, estado, departamento, ubicacion)}",
        reply_markup=construir_menu(update.effective_user.id),
    )
    context.user_data.clear()
    return ConversationHandler.END


async def editar_guardar_valor_estado(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    nuevo_valor = query.data.split(":")[1]
    expediente_id = context.user_data.get("editar_id")
    actualizar_campo_expediente(expediente_id, "estado", nuevo_valor)
    row = obtener_expediente_por_id(expediente_id)
    _, nombre, carnet, estado, departamento, ubicacion = row
    await query.edit_message_text(
        f"✅ Actualizado.\n\n{formatear_expediente(nombre, carnet, estado, departamento, ubicacion)}"
    )
    context.user_data.clear()
    return ConversationHandler.END


async def editar_guardar_valor_departamento(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    indice = int(query.data.split(":")[1])
    nuevo_valor = DEPARTAMENTOS[indice]
    expediente_id = context.user_data.get("editar_id")
    actualizar_campo_expediente(expediente_id, "departamento", nuevo_valor)
    row = obtener_expediente_por_id(expediente_id)
    _, nombre, carnet, estado, departamento, ubicacion = row
    await query.edit_message_text(
        f"✅ Actualizado.\n\n{formatear_expediente(nombre, carnet, estado, departamento, ubicacion)}"
    )
    context.user_data.clear()
    return ConversationHandler.END


async def manejar_texto_libre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Intenta interpretar un mensaje de texto libre como una búsqueda natural."""
    if not await verificar_acceso(update):
        return

    termino = detectar_busqueda_natural(update.message.text)
    if not termino:
        await update.message.reply_text(
            f"No te entendí. Puedes escribirme algo como \"{NOMBRE_BOT}, ¿dónde está el expediente de Ali?\", "
            "o usar el menú de abajo.",
            reply_markup=construir_menu(update.effective_user.id),
        )
        return

    await responder_busqueda(update, termino)


# ---------- Main ----------

def main():
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    if not token:
        raise RuntimeError(
            "Falta la variable de entorno TELEGRAM_BOT_TOKEN. "
            "Configúrala con el token que te dio @BotFather."
        )

    init_db()

    app = ApplicationBuilder().token(token).build()

    texto_no_comando = filters.TEXT & ~filters.COMMAND

    conv_buscar = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{BTN_BUSCAR}$"), buscar_start)],
        states={ESPERANDO_BUSQUEDA: [MessageHandler(texto_no_comando, buscar_recibir)]},
        fallbacks=[
            MessageHandler(filters.Regex(f"^{BTN_CANCELAR}$"), cancelar),
            CommandHandler("cancelar", cancelar),
        ],
    )

    conv_agregar = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{BTN_AGREGAR}$"), agregar_start)],
        states={
            NOMBRE: [MessageHandler(texto_no_comando, agregar_nombre)],
            CARNET: [MessageHandler(texto_no_comando, agregar_carnet)],
            DECIDIR_DUPLICADO: [CallbackQueryHandler(decidir_duplicado, pattern=r"^(usar_existente|crear_nuevo)$")],
            ESTADO: [CallbackQueryHandler(agregar_estado_boton, pattern=r"^estado:")],
            DEPARTAMENTO: [CallbackQueryHandler(agregar_departamento_boton, pattern=r"^depto:")],
            UBICACION: [MessageHandler(texto_no_comando, agregar_ubicacion)],
        },
        fallbacks=[
            MessageHandler(filters.Regex(f"^{BTN_CANCELAR}$"), cancelar),
            CommandHandler("cancelar", cancelar),
        ],
    )

    conv_baja = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{BTN_BAJA}$"), baja_start)],
        states={ESPERANDO_BAJA: [MessageHandler(texto_no_comando, baja_recibir)]},
        fallbacks=[
            MessageHandler(filters.Regex(f"^{BTN_CANCELAR}$"), cancelar),
            CommandHandler("cancelar", cancelar),
        ],
    )

    conv_retirar = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{BTN_RETIRAR}$"), retirar_start)],
        states={ESPERANDO_RETIRAR: [MessageHandler(texto_no_comando, retirar_recibir)]},
        fallbacks=[
            MessageHandler(filters.Regex(f"^{BTN_CANCELAR}$"), cancelar),
            CommandHandler("cancelar", cancelar),
        ],
    )

    conv_eliminar = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{BTN_ELIMINAR}$"), eliminar_start)],
        states={ESPERANDO_ELIMINAR: [MessageHandler(texto_no_comando, eliminar_recibir)]},
        fallbacks=[
            MessageHandler(filters.Regex(f"^{BTN_CANCELAR}$"), cancelar),
            CommandHandler("cancelar", cancelar),
        ],
    )

    conv_editar = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex(f"^{BTN_EDITAR}$"), editar_start)],
        states={
            ESPERANDO_BUSQUEDA_EDITAR: [
                MessageHandler(texto_no_comando, editar_buscar_recibir),
                CallbackQueryHandler(editar_elegir_callback, pattern=r"^editar_elegir:\d+$"),
            ],
            ELIGIENDO_CAMPO: [CallbackQueryHandler(editar_elegir_campo, pattern=r"^campo:")],
            ESPERANDO_VALOR_CAMPO: [
                MessageHandler(texto_no_comando, editar_guardar_valor_texto),
                CallbackQueryHandler(editar_guardar_valor_estado, pattern=r"^nuevoestado:"),
                CallbackQueryHandler(editar_guardar_valor_departamento, pattern=r"^depto:"),
            ],
        },
        fallbacks=[
            MessageHandler(filters.Regex(f"^{BTN_CANCELAR}$"), cancelar),
            CommandHandler("cancelar", cancelar),
        ],
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("ayuda", ayuda))
    app.add_handler(CommandHandler("rol", cambiar_rol))
    app.add_handler(CommandHandler("revocar", revocar))

    app.add_handler(conv_buscar)
    app.add_handler(conv_agregar)
    app.add_handler(conv_baja)
    app.add_handler(conv_retirar)
    app.add_handler(conv_eliminar)
    app.add_handler(conv_editar)

    app.add_handler(MessageHandler(filters.Regex(f"^{BTN_LISTA}$"), mostrar_lista))
    app.add_handler(MessageHandler(filters.Regex(f"^{BTN_EXPORTAR}$"), exportar))
    app.add_handler(MessageHandler(filters.Regex(f"^{BTN_USUARIOS}$"), mostrar_usuarios))
    app.add_handler(MessageHandler(filters.Regex(f"^{BTN_AYUDA}$"), ayuda))

    app.add_handler(
        CallbackQueryHandler(manejar_aprobacion, pattern=r"^(aprobar|rechazar):(editor|lector|-):\d+$")
    )
    app.add_handler(
        CallbackQueryHandler(manejar_confirmacion_eliminar, pattern=r"^(confirmar_eliminar:\d+|cancelar_eliminar)$")
    )
    app.add_handler(CallbackQueryHandler(manejar_ver_expediente, pattern=r"^ver_exp:\d+$"))

    # Este handler va al final: solo captura texto que no fue un botón
    # ni parte de una conversación activa, e intenta interpretarlo como
    # una búsqueda en lenguaje natural (ej. "Lucas, ¿dónde está Ali?").
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, manejar_texto_libre))

    logger.info("Bot iniciado. Esperando mensajes...")
    app.run_polling()


if __name__ == "__main__":
    main()
